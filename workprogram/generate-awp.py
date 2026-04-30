#!/usr/bin/env python3
"""
NACSA CoP Core Audit Work Program (AWP) Generator — v5.0
Generates: NACSA-CoP-Core-WorkProgram.xlsx

Rebuilt to follow IESP 14-column pattern:
- 14-column layout: CoP Ref | AWP Ref | Level | Control Domain | Sub-Procedure |
  Assessment Procedures | Expected Evidence | Method | Procedures Performed |
  Evidence Obtained | Evidence Ref | Observations | Conclusion | Recommendations
- Domain-level conclusions with sub-item observations
- Context rows (WHY / WHAT GOOD LOOKS LIKE / KEY RISK IF ABSENT)
- Tier 2 rows visually distinct (italic, grey fill)
- SYSTEM-level rows pre-populated per NCII system placeholder
- COUNTIF scoring dashboard
- 25 sheets: 8 standard + 17 domain worksheets

Data sources:
- controls/library.json v4.0 (126 controls, 228 requirements from CoP clauses)
- controls/domains.json v3.0 (18 domains, 17 with controls)
- AWP-Requirements-v2/awp-requirements.json (42 directive-only requirements)
"""

import json
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.normpath(os.path.join(SCRIPT_DIR, ".."))

# ── Styles (matching IESP pattern) ────────────────────────────────
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
SECTION_FONT = Font(name='Calibri', bold=True, size=11, color='2F5496')
SECTION_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
DOMAIN_FONT = Font(name='Calibri', bold=True, size=10)
DOMAIN_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
BODY_FONT = Font(name='Calibri', size=10)
CONTEXT_FONT = Font(name='Calibri', size=10, italic=True)
CONTEXT_FILL = PatternFill(start_color='E8D5E8', end_color='E8D5E8', fill_type='solid')
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='2F5496')
SUBTITLE_FONT = Font(name='Calibri', bold=True, size=12, color='2F5496')
LABEL_FONT = Font(name='Calibri', bold=True, size=10)
VALUE_FONT = Font(name='Calibri', size=10)
SMALL_FONT = Font(name='Calibri', size=9, italic=True)
LEGAL_FILL = PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid')
LEGAL_HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

# Tier 2 styling
TIER2_FONT = Font(name='Calibri', size=10, italic=True)
TIER2_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

# Scoring fills
COMPLIANT_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
PARTIAL_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
NON_COMPLIANT_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
NA_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
CONCLUSION_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
LIGHT_GREY = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
ORG_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
SYS_FILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

# Directive sheet
DIRECTIVE_FILL = PatternFill(start_color='E8D5E8', end_color='E8D5E8', fill_type='solid')
DIRECTIVE_HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
DIRECTIVE_HEADER_FILL = PatternFill(start_color='5B2C6F', end_color='5B2C6F', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))
WRAP_ALIGN = Alignment(wrap_text=True, vertical='top')
CENTER_ALIGN = Alignment(horizontal='center', vertical='top', wrap_text=True)

# 14-column layout
COL_WIDTHS = {
    'A': 14,   # CoP Ref
    'B': 12,   # AWP Ref
    'C': 11,   # Level
    'D': 24,   # Control Domain
    'E': 30,   # Sub-Procedure
    'F': 50,   # Assessment Procedures
    'G': 35,   # Expected Evidence
    'H': 14,   # Method
    'I': 30,   # Procedures Performed
    'J': 25,   # Evidence Obtained
    'K': 12,   # Evidence Ref
    'L': 30,   # Observations
    'M': 16,   # Conclusion
    'N': 30,   # Recommendations
}

ASSESSMENT_HEADERS = [
    'CoP Ref', 'AWP Ref', 'Level', 'Control Domain', 'Sub-Procedure',
    'Assessment Procedures', 'Expected Evidence', 'Method',
    'Procedures Performed', 'Evidence Obtained', 'Evidence Ref',
    'Observations', 'Conclusion', 'Recommendations'
]
NUM_COLS = len(ASSESSMENT_HEADERS)  # 14


# ── Data Loading ──────────────────────────────────────────────────

def load_json(path):
    with open(path, "r") as f:
        return json.load(f)


# Load library
library_data = load_json(os.path.join(REPO_ROOT, "controls", "library.json"))
domains_data = load_json(os.path.join(REPO_ROOT, "controls", "domains.json"))
CONTROLS_LIST = library_data.get("controls", [])

# Parse domains
COP_DOMAINS = []
for key, val in domains_data.items():
    if key == "_meta":
        continue
    COP_DOMAINS.append({
        "slug": key,
        "name": val["name"],
        "copSection": val["copSection"],
        "ncsbCategory": val.get("ncsbCategory"),
        "ncsbDomain": val.get("ncsbDomain"),
        "description": val["description"],
        "relatedSections": val.get("relatedSections", []),
    })
COP_DOMAINS.sort(key=lambda d: float(d["copSection"]))

# Group controls by domain
CONTROLS_BY_DOMAIN = {}
for c in CONTROLS_LIST:
    d = c["domain"]
    if d not in CONTROLS_BY_DOMAIN:
        CONTROLS_BY_DOMAIN[d] = []
    CONTROLS_BY_DOMAIN[d].append(c)

# Load directive requirements
AWP_REQ_PATH = os.path.expanduser(
    "~/claude/tech-audit/NACSA/Research/AWP-Requirements-v2/awp-requirements.json"
)
DIRECTIVE_REQUIREMENTS = []
if os.path.exists(AWP_REQ_PATH):
    awp_req_data = load_json(AWP_REQ_PATH)
    DIRECTIVE_REQUIREMENTS = [
        r for r in awp_req_data["requirements"]
        if r.get("ceDirective") is not None
    ]
else:
    print(f"WARNING: Directive requirements not found at {AWP_REQ_PATH}")

# Placeholder NCII systems
NCII_SYSTEMS = [
    {"id": "NCII-SYS-001", "name": "[System 1]", "description": "[Description]"},
    {"id": "NCII-SYS-002", "name": "[System 2]", "description": "[Description]"},
    {"id": "NCII-SYS-003", "name": "[System 3]", "description": "[Description]"},
]


# ── Helper Functions ──────────────────────────────────────────────

def apply_col_widths(ws):
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


def write_header_row(ws, row):
    for col, header in enumerate(ASSESSMENT_HEADERS, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def write_section_row(ws, row, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.border = THIN_BORDER
    for col in range(2, NUM_COLS + 1):
        ws.cell(row=row, column=col).fill = SECTION_FILL
        ws.cell(row=row, column=col).border = THIN_BORDER


def write_context_row(ws, row, context):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
    cell = ws.cell(row=row, column=1, value=context)
    cell.font = CONTEXT_FONT
    cell.fill = CONTEXT_FILL
    cell.alignment = WRAP_ALIGN
    cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 60
    for col in range(2, NUM_COLS + 1):
        ws.cell(row=row, column=col).fill = CONTEXT_FILL
        ws.cell(row=row, column=col).border = THIN_BORDER


def write_domain_row(ws, row, cop_ref, awp_ref, level, control_domain, summary):
    data = [cop_ref, awp_ref, level, control_domain, summary,
            '', '', '', '', '', '', '', '', '']
    for col, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = DOMAIN_FONT
        cell.fill = DOMAIN_FILL
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER
    # Highlight conclusion and recommendation columns
    ws.cell(row=row, column=13).fill = CONCLUSION_FILL
    ws.cell(row=row, column=14).fill = CONCLUSION_FILL


def write_subitem_row(ws, row, cop_ref, awp_ref, sub_proc, procedures, evidence,
                      method='Inspection', is_tier2=False, alt_row=False):
    data = [cop_ref, awp_ref, '', '', sub_proc, procedures, evidence, method,
            '', '', '', '', '', '']
    font = TIER2_FONT if is_tier2 else BODY_FONT
    fill = TIER2_FILL if is_tier2 else (LIGHT_GREY if alt_row else None)
    for col, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def write_row(ws, row, values, font=None, fill=None, alignment=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font or BODY_FONT
        if fill:
            cell.fill = fill
        cell.alignment = alignment or WRAP_ALIGN
        cell.border = THIN_BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Statutory Data ────────────────────────────────────────────────

STATUTORY_OBLIGATIONS = [
    ("s17", "NCII Entity Designation", "Respond to designation notice within prescribed timeframe", "s18(1)(a)", "RM500,000 / 10 years", "Yes (s55)", "Entity registration, designation acknowledgement"),
    ("s18", "General NCII Duties", "Implement and maintain cyber security measures to protect NCII", "s18(2)", "RM500,000 / 10 years", "Yes (s55)", "Comprehensive controls across all CoP domains"),
    ("s20", "Information Provision", "Provide information/documents when required by CE", "s20(3)", "RM200,000 / 3 years", "Yes (s55)", "Response records, submission logs"),
    ("s20(4)", "False/Misleading Information", "Ensure accuracy of all information provided to CE", "s20(4)", "RM300,000 / 5 years", "No", "Accuracy verification process"),
    ("s21", "Risk Assessment", "Conduct annual RA by CE-approved persons, submit within 30 days", "s21(3)", "RM500,000 / 10 years", "Yes (s55)", "RA report, assessor approval, submission confirmation"),
    ("s22", "Code of Practice Compliance", "Comply with sector-specific CoP approved by CE", "s22(2)", "RM200,000 / 3 years", "Yes (s55)", "CoP compliance assessment, gap tracker"),
    ("s23", "Cyber Security Audit", "Conduct biennial audit by CE-approved auditor, submit within 30 days", "s23(3)", "RM500,000 / 10 years", "Yes (s55)", "Audit report, auditor approval, submission confirmation"),
    ("s24", "Standards Compliance", "Comply with cyber security standards set by CE", "s24(2)", "RM200,000 / 3 years", "Yes (s55)", "Standards compliance evidence"),
    ("s25", "Threat Information Sharing", "Inform CE of threats that may affect other NCII", "s25(2)", "RM200,000 / 3 years", "Yes (s55)", "Threat sharing procedure, submission records"),
    ("s26", "Incident Notification", "Notify CE within 6 hours of significant-impact incident", "s26(3)", "RM500,000 / 10 years + RM5,000/day", "Yes (s55)", "IRP, NC4 portal access, notification records"),
    ("s26(4)", "False Incident Notification", "Ensure accuracy of incident notifications", "s26(4)", "RM300,000 / 5 years", "No", "Notification accuracy verification"),
]

S52_ASSESSMENT = [
    ("Board-level cyber security governance", "Does the board have formal cyber security governance procedures? Is there a dedicated committee?", "Board charter, committee ToR, meeting minutes (12 months)"),
    ("Active vs passive governance", "Do board minutes show active engagement (questioning, directing) or passive receipt of reports?", "Board minutes analysis, decision records"),
    ("CISO/ICTSO board reporting", "Does the CISO have direct board access? How frequently are briefings conducted?", "CISO appointment letter, reporting schedule, briefing records"),
    ("Delegation with accountability", "Is cyber security delegation documented with retained board accountability?", "Delegation instruments, performance monitoring records"),
    ("Response to known risks", "When audit or RA identified gaps, did the board direct remediation?", "Risk register, remediation tracker, board action items"),
    ("Director training on Act 854", "Have directors been trained on their personal liability under s52?", "Training records, signed acknowledgements"),
    ("D&O insurance coverage", "Does D&O coverage extend to Act 854 criminal proceedings? Are limits adequate (10x multiplier)?", "D&O policy, coverage analysis"),
    ("Due diligence documentation", "Could each director demonstrate 'all due diligence' per s52(2) if required?", "Composite assessment of above evidence"),
]

PENALTY_MATRIX = [
    ("s18(2)", "Failure to implement NCII duties", "RM500,000", "10 years", "RM5,000,000", "Yes", "50% = RM250,000"),
    ("s20(3)", "Failure to provide information", "RM200,000", "3 years", "RM2,000,000", "Yes", "50% = RM100,000"),
    ("s20(4)", "False/misleading information", "RM300,000", "5 years", "RM3,000,000", "No", "N/A"),
    ("s21(3)", "Failure to conduct risk assessment", "RM500,000", "10 years", "RM5,000,000", "Yes", "50% = RM250,000"),
    ("s22(2)", "Failure to comply with CoP", "RM200,000", "3 years", "RM2,000,000", "Yes", "50% = RM100,000"),
    ("s23(3)", "Failure to conduct audit", "RM500,000", "10 years", "RM5,000,000", "Yes", "50% = RM250,000"),
    ("s24(2)", "Failure to comply with standards", "RM200,000", "3 years", "RM2,000,000", "Yes", "50% = RM100,000"),
    ("s25(2)", "Failure to share threat information", "RM200,000", "3 years", "RM2,000,000", "Yes", "50% = RM100,000"),
    ("s26(3)", "Failure to notify incident (6 hours)", "RM500,000 + RM5,000/day", "10 years", "RM5,000,000 + RM50,000/day", "Yes", "50% of max fine"),
    ("s26(4)", "False incident notification", "RM300,000", "5 years", "RM3,000,000", "No", "N/A"),
    ("s44", "Obstruction of authorised officer", "RM300,000", "5 years", "RM3,000,000", "Conditional", "If no force/threats"),
    ("s49", "Breach of secrecy", "RM500,000", "10 years", "RM5,000,000", "No", "N/A"),
]

DIRECTIVE_NAMES = {
    "01": "Directive 01: Incident Notification & Authorised Persons",
    "02": "Directive 02: CSSP Licensing",
    "04": "Directive 04: Self-Assessment",
    "05": "Directive 05: Risk Assessment",
    "08": "Directive 08: Audit Process",
    "09": "Directive 09: PQC Migration",
    "10": "Directive 10: Crisis Management",
}


# ── Context Data (WHY / WHAT GOOD LOOKS LIKE / KEY RISK IF ABSENT) ──

DOMAIN_CONTEXT = {
    "governance-leadership": (
        "WHY: Governance sets the tone from the top. Act 854 s52 creates personal criminal liability for directors/officers "
        "who consent, connive, or neglect cyber security duties. Without board-level ownership, cyber security becomes an IT "
        "afterthought rather than a business priority.\n"
        "WHAT GOOD LOOKS LIKE: Dedicated governance committee with ToR, regular board briefings from CISO/ICTSO, documented "
        "delegation with retained accountability, directors trained on s52 obligations.\n"
        "KEY RISK IF ABSENT: Directors personally liable under s52 with reverse burden of proof. Entity lacks strategic "
        "direction for cyber security investment and prioritisation."
    ),
    "cyber-security-policy": (
        "WHY: The cyber security policy is the foundational document that cascades into all other controls. It sets the "
        "organisation's commitment, scope, and approach to protecting NCII assets.\n"
        "WHAT GOOD LOOKS LIKE: Approved policy aligned to ISO 27001/NIST, communicated to all personnel, reviewed annually, "
        "with measurable SMART objectives tracked and reported to governance committee.\n"
        "KEY RISK IF ABSENT: No documented basis for control implementation. Staff unclear on expectations. CoP s22 "
        "non-compliance exposure."
    ),
    "organisational-development": (
        "WHY: People are both the strongest and weakest link. Adequate staffing, budget, authority contacts (NC4/NACSA), "
        "and security awareness ensure the organisation can detect, prevent, and respond to threats.\n"
        "WHAT GOOD LOOKS LIKE: Qualified CISO/ICTSO with defined team, dedicated budget, current NC4/NACSA contacts, "
        "annual training plan with measured effectiveness, phishing simulations.\n"
        "KEY RISK IF ABSENT: Understaffed security function, untrained users susceptible to social engineering, inability "
        "to contact authorities during incidents."
    ),
    "cyber-security-assurance": (
        "WHY: Assurance activities (acceptable use, audit, supplier security) verify that controls work and third parties "
        "don't introduce risk. Biennial audit is mandatory under s23; supplier risk management protects the supply chain.\n"
        "WHAT GOOD LOOKS LIKE: Enforced IT AUP with signed acknowledgements, biennial audit completed and submitted to CE "
        "within 30 days, comprehensive supplier register with CSSP licensing verification, remediation of prior findings.\n"
        "KEY RISK IF ABSENT: Uncontrolled user behaviour, missed audit deadlines (s23 penalty), supply chain compromise."
    ),
    "resource-allocation": (
        "WHY: You cannot protect what you don't know exists. Asset management and capacity planning ensure NCII systems "
        "are identified, classified, and adequately resourced. Act 854 s17 designation depends on knowing what qualifies.\n"
        "WHAT GOOD LOOKS LIKE: Complete CMDB/asset register with classification, assigned ownership, quarterly maintenance, "
        "capacity monitoring with alerts, annual forecasting, performance testing before go-live.\n"
        "KEY RISK IF ABSENT: Unmanaged shadow assets, unpatched systems, capacity failures under load, NCII not properly "
        "reported to sector lead."
    ),
    "risk-management": (
        "WHY: Risk assessment is mandatory annually under P.U.(A) 219/2024 (s21). It drives control prioritisation — "
        "without it, the entity is guessing which threats matter most.\n"
        "WHAT GOOD LOOKS LIKE: Documented RA methodology aligned to ISO 27005/NIST 800-30, annual RA covering all NCII, "
        "submitted to CE within 30 days, risk treatment plan approved by governance committee, residual risk accepted.\n"
        "KEY RISK IF ABSENT: s21 non-compliance (RM500,000 / 10 years), uninformed control investment, unidentified "
        "critical risks."
    ),
    "operational-efficiency": (
        "WHY: Change management prevents unauthorised modifications that could compromise NCII. Uncontrolled changes are a "
        "leading cause of outages and security incidents.\n"
        "WHAT GOOD LOOKS LIKE: Documented change procedure covering request, risk assessment, approval, testing, "
        "implementation, and review. Emergency changes with retrospective approval. Annual review of process.\n"
        "KEY RISK IF ABSENT: Unauthorised changes causing security gaps or outages, inability to trace changes back to "
        "approved requests, audit trail gaps."
    ),
    "data-security": (
        "WHY: Data is the ultimate target. DLP, privacy (PDPA Act 709), application security, and secure configuration "
        "form layers of protection around the data that NCII systems process.\n"
        "WHAT GOOD LOOKS LIKE: DLP deployed and integrated with SIEM, PDPA-aligned privacy procedures, OWASP-based "
        "secure coding standards, CIS/STIG hardening baselines applied to all systems, FIM monitoring critical files.\n"
        "KEY RISK IF ABSENT: Data exfiltration, PDPA non-compliance, application vulnerabilities, misconfigured systems "
        "as easy targets."
    ),
    "contractual-management": (
        "WHY: NDAs protect sensitive information shared during NCII operations. Without them, confidential data disclosed "
        "to third parties has no legal protection.\n"
        "WHAT GOOD LOOKS LIKE: NDA templates covering all party types, signed before information exchange, compliance "
        "monitoring with periodic review, breaches addressed through contractual remedies.\n"
        "KEY RISK IF ABSENT: s49 breach of secrecy exposure, unprotected disclosure of NCII-related information."
    ),
    "physical-security": (
        "WHY: Physical compromise bypasses all logical controls. Physical security protects NCII infrastructure from "
        "unauthorised access, theft, and environmental threats.\n"
        "WHAT GOOD LOOKS LIKE: Multi-layered access controls (badge/biometric), CCTV with 90+ day retention, visitor "
        "management, environmental monitoring, defined security perimeters with zoning.\n"
        "KEY RISK IF ABSENT: Unauthorised physical access to servers/network equipment, environmental damage to NCII "
        "infrastructure."
    ),
    "system-network-security": (
        "WHY: Network and system security controls are the technical backbone of NCII protection. Firewalls, IDS/IPS, "
        "NAC, and malware prevention defend against external and internal threats.\n"
        "WHAT GOOD LOOKS LIKE: Documented network security policy with logical/physical diagrams, firewall rules following "
        "least-privilege, IDS/IPS at critical boundaries, NAC enforcing device compliance, malware prevention integrated "
        "with SIEM.\n"
        "KEY RISK IF ABSENT: Network compromise providing access to NCII systems, undetected lateral movement, malware "
        "propagation across segments."
    ),
    "access-control": (
        "WHY: Access control ensures only authorised individuals access NCII systems with appropriate privileges. IAM, "
        "SoD, and password management are foundational controls.\n"
        "WHAT GOOD LOOKS LIKE: IAM with MFA enforced on all NCII access, documented SoD matrix, password management "
        "system aligned to NIST 800-63B, regular access reviews with revocation evidence.\n"
        "KEY RISK IF ABSENT: Excessive privileges, shared accounts, compromised credentials, insider threat, segregation "
        "of duties failures enabling fraud."
    ),
    "technical-vulnerability": (
        "WHY: Vulnerabilities are the attack surface. Without systematic identification, assessment, and remediation, "
        "known exploitable weaknesses persist in NCII systems.\n"
        "WHAT GOOD LOOKS LIKE: Documented VMP with SLA-based remediation timelines, regular vulnerability scans, "
        "penetration testing, tracked remediation with risk acceptance for exceptions.\n"
        "KEY RISK IF ABSENT: Unpatched systems exploited by known attack vectors, extended exposure windows, no "
        "visibility into security posture."
    ),
    "event-management": (
        "WHY: Security monitoring is the detective control that identifies threats in progress. Without monitoring, "
        "breaches go undetected — the average dwell time for undetected breaches is 200+ days.\n"
        "WHAT GOOD LOOKS LIKE: Documented SMP with defined scope and tools (IDPS/SIEM/EDR), threat intelligence "
        "integration, MTTD/MTTR KPIs tracked and reported, 24/7 coverage.\n"
        "KEY RISK IF ABSENT: Undetected intrusions, extended dwell time, inability to meet 6-hour incident notification "
        "requirement (s26)."
    ),
    "incident-management": (
        "WHY: Incident management is critical — Act 854 requires 6-hour notification to CE for significant incidents "
        "(s26). The CSIRT, NC4 portal access, and tested playbooks determine whether the entity can respond in time.\n"
        "WHAT GOOD LOOKS LIKE: CSIMP covering Plan/Detect/Respond/Update, designated CSIRT with 3 authorised persons, "
        "NC4 portal operational, 6-hour notification tabletop tested, forensic capability, post-incident review.\n"
        "KEY RISK IF ABSENT: Missed 6-hour notification window (s26 penalty with daily compounding RM5,000), uncoordinated "
        "response, evidence destruction, prolonged recovery."
    ),
    "business-continuity": (
        "WHY: Business continuity ensures NCII services recover from disruptions. BCP with cyber scenarios, tested DR, "
        "and adequate backup protect against extended outages.\n"
        "WHAT GOOD LOOKS LIKE: BCP with BIA defining RTO/RPO per NCII system, crisis management procedures, annual "
        "DR failover testing with security validation, backup integrity verified through restore tests.\n"
        "KEY RISK IF ABSENT: Extended NCII service disruption, untested recovery procedures that fail under real "
        "conditions, data loss from backup failures."
    ),
    "sector-specific": (
        "WHY: NCII entities operate in specific sectors with unique threat profiles and regulatory requirements. "
        "Sector-specific CoPs address OT security (energy, transport), e-commerce security (banking), and other "
        "sector-unique controls.\n"
        "WHAT GOOD LOOKS LIKE: Entity has obtained its sector-specific CoP from NACSA/sector lead, performed "
        "self-assessment against sector requirements, addressed OT/ICS security if applicable.\n"
        "KEY RISK IF ABSENT: Sector-specific attack vectors unaddressed, non-compliance with sector CoP requirements."
    ),
}


# ── Assessment Procedures and Evidence by Control ─────────────────

def get_assessment_procedures(control, req):
    """Generate assessment procedures based on the control domain and clause content."""
    clause = req.get("clauseNumber", "")
    text = req.get("text", "").lower()
    subsection = control.get("copSubsection", "")
    domain = control.get("domain", "")

    # Generate contextual procedures based on the requirement text and domain
    procedures = []
    evidence = []

    # Policy-related
    if any(w in text for w in ["policy", "establish", "develop"]) and "shall" in text.lower():
        procedures.extend([
            "1. Request the current approved document.",
            "2. Verify it is approved by the governance committee (check signature/approval record).",
            "3. Confirm it covers all required elements specified in the CoP clause.",
            "4. Check it is current (approved within the last 12 months or review cycle)."
        ])
        evidence.extend(["Approved policy/procedure document", "Committee approval record", "Version history"])

    # Review-related
    elif "review" in text and ("annual" in text or "annually" in text):
        procedures.extend([
            "1. Request evidence of the most recent review (minutes, review report).",
            "2. Verify the review was conducted within the required timeframe.",
            "3. Confirm review addressed all required elements per the CoP clause.",
            "4. Check that review outcomes were acted upon (changes made, approvals obtained)."
        ])
        evidence.extend(["Review report/minutes", "Updated document post-review", "Approval of changes"])

    # Communication-related
    elif "communicat" in text or "disseminat" in text:
        procedures.extend([
            "1. Request evidence of communication to relevant stakeholders.",
            "2. Verify communication method is appropriate (training, intranet, briefings).",
            "3. Sample 10 staff — confirm acknowledgement/receipt of communication.",
            "4. Confirm communication is current (reflects latest version)."
        ])
        evidence.extend(["Communication records", "Staff acknowledgement samples", "Intranet/portal screenshots"])

    # Training-related
    elif "training" in text or "awareness" in text:
        procedures.extend([
            "1. Request training plan and records for the assessment period.",
            "2. Verify training content covers required topics per CoP clause.",
            "3. Check completion rates (target: 100% for mandatory training).",
            "4. Review effectiveness evaluation results (post-training assessments)."
        ])
        evidence.extend(["Training plan", "Completion records/LMS reports", "Assessment results", "Training materials"])

    # Audit/assessment-related
    elif "audit" in text or "assessment" in text:
        procedures.extend([
            "1. Request the most recent audit/assessment report.",
            "2. Verify it was conducted by an approved/qualified party.",
            "3. Confirm submission to CE NACSA within required timeframe.",
            "4. Check that findings are tracked with remediation plans."
        ])
        evidence.extend(["Audit/assessment report", "Assessor qualifications/approval", "CE submission confirmation", "Remediation tracker"])

    # Risk-related
    elif "risk" in text and ("assess" in text or "identif" in text or "evaluat" in text):
        procedures.extend([
            "1. Request the risk assessment methodology and most recent report.",
            "2. Verify scope includes all NCII systems.",
            "3. Confirm risk identification, analysis, and evaluation steps documented.",
            "4. Check risk register is current with treatment plans."
        ])
        evidence.extend(["Risk assessment methodology", "Risk assessment report", "Risk register", "Risk treatment plan"])

    # Technical controls - firewall/network
    elif any(w in text for w in ["firewall", "network security", "ids", "ips", "nac"]):
        procedures.extend([
            "1. Request configuration documentation for the control.",
            "2. Verify deployment covers all NCII network segments.",
            "3. Review rule base/configuration for least-privilege and best practice.",
            "4. Check for regular review and update of configurations."
        ])
        evidence.extend(["Configuration documentation", "Rule base export", "Review records", "Network diagram"])

    # Access control / IAM
    elif any(w in text for w in ["access control", "access management", "mfa", "multi-factor", "password"]):
        procedures.extend([
            "1. Request the IAM/access control procedure.",
            "2. Verify MFA deployment on NCII system access.",
            "3. Sample 15 user accounts — verify access rights match job roles.",
            "4. Check access review records (quarterly for privileged, semi-annually for standard)."
        ])
        evidence.extend(["IAM procedure", "MFA deployment evidence", "User access sample", "Access review records"])

    # Monitoring/SIEM
    elif any(w in text for w in ["monitor", "siem", "log", "detect"]):
        procedures.extend([
            "1. Request the monitoring procedure/SMP.",
            "2. Verify monitoring tools deployed (SIEM, IDPS, EDR).",
            "3. Check log source coverage includes all NCII systems.",
            "4. Review MTTD/MTTR KPI performance for the assessment period."
        ])
        evidence.extend(["Monitoring procedure", "Tool deployment evidence", "Log source inventory", "KPI dashboard/reports"])

    # Incident-related
    elif "incident" in text and ("notif" in text or "report" in text or "respond" in text):
        procedures.extend([
            "1. Request the incident management procedure (CSIMP).",
            "2. Verify CSIRT membership and authorised person designations.",
            "3. Confirm NC4 portal access and 6-hour notification capability.",
            "4. Review incident log and tabletop exercise results."
        ])
        evidence.extend(["CSIMP document", "CSIRT charter", "NC4 portal evidence", "Incident log", "Tabletop results"])

    # BCP/DR
    elif any(w in text for w in ["business continuity", "disaster recovery", "backup", "bcp", "drp"]):
        procedures.extend([
            "1. Request BCP/DRP documentation.",
            "2. Verify BIA with defined RTO/RPO per NCII system.",
            "3. Check most recent DR failover test results.",
            "4. Verify backup procedures and restore test evidence."
        ])
        evidence.extend(["BCP/DRP document", "BIA with RTO/RPO", "DR test results", "Backup/restore test records"])

    # Supplier/vendor
    elif any(w in text for w in ["supplier", "vendor", "third-party", "service provider", "supply chain"]):
        procedures.extend([
            "1. Request supplier/vendor register.",
            "2. Verify CSSP licensing for applicable providers (P.U.(A) 221/2024).",
            "3. Check contractual security requirements in supplier agreements.",
            "4. Review supplier risk assessments and compliance monitoring."
        ])
        evidence.extend(["Supplier register", "CSSP licence records", "Sample supplier agreements", "Risk assessment records"])

    # DLP / Data protection
    elif any(w in text for w in ["data leakage", "dlp", "data privacy", "personal data", "pdpa"]):
        procedures.extend([
            "1. Request DLP/data protection documentation.",
            "2. Verify DLP solution deployment scope covers NCII data.",
            "3. Check data classification applied to NCII assets.",
            "4. Review DLP alerts and incidents for the assessment period."
        ])
        evidence.extend(["DLP procedure", "DLP deployment evidence", "Data classification records", "DLP alert logs"])

    # Asset management
    elif any(w in text for w in ["asset", "inventory", "cmdb"]):
        procedures.extend([
            "1. Request the asset register/CMDB extract.",
            "2. Verify all NCII systems are included with required attributes.",
            "3. Sample 15 assets — confirm data accuracy (owner, classification, status).",
            "4. Check maintenance plan and review records."
        ])
        evidence.extend(["Asset register/CMDB", "Asset classification records", "Maintenance plan", "Review records"])

    # Vulnerability management
    elif any(w in text for w in ["vulnerabilit", "scan", "penetration", "patch"]):
        procedures.extend([
            "1. Request vulnerability management procedure.",
            "2. Review latest vulnerability scan results for NCII systems.",
            "3. Check remediation timelines against SLAs.",
            "4. Verify penetration test reports and remediation of findings."
        ])
        evidence.extend(["VMP document", "Scan reports", "Remediation tracker", "Penetration test reports"])

    # Change management
    elif "change" in text and ("management" in text or "control" in text or "process" in text):
        procedures.extend([
            "1. Request change management procedure.",
            "2. Sample 10 changes — verify full lifecycle (request, risk assessment, approval, test, implementation).",
            "3. Check for emergency change records with retrospective approval.",
            "4. Review annual change process review."
        ])
        evidence.extend(["Change management procedure", "10 sampled change records", "Emergency change records", "Process review records"])

    # Physical security
    elif any(w in text for w in ["physical", "cctv", "perimeter", "visitor"]):
        procedures.extend([
            "1. Request physical security procedure.",
            "2. Verify multi-layered access controls for sensitive areas.",
            "3. Check CCTV coverage map and retention period (minimum 90 days).",
            "4. Review visitor management logs and emergency response procedures."
        ])
        evidence.extend(["Physical security procedure", "Access control system config", "CCTV coverage map", "Visitor logs"])

    # NDA
    elif "nda" in text or "non-disclosure" in text or "confidential" in text:
        procedures.extend([
            "1. Request NDA templates and signed NDA samples.",
            "2. Verify NDA templates cover required parties.",
            "3. Sample 10 NDAs — confirm signed before information exchange.",
            "4. Check NDA compliance monitoring records."
        ])
        evidence.extend(["NDA templates", "Signed NDA samples", "Compliance monitoring records"])

    # Capacity management
    elif any(w in text for w in ["capacity", "performance test", "resource utilisation"]):
        procedures.extend([
            "1. Request capacity management plan or monitoring procedure.",
            "2. Verify resource utilisation monitoring covers CPU, memory, network, disk.",
            "3. Check alerting thresholds and escalation procedures.",
            "4. Review forecasting data and performance test results."
        ])
        evidence.extend(["Capacity management plan", "Monitoring dashboards", "Alert configuration", "Performance test results"])

    # Secure coding / application security
    elif any(w in text for w in ["secure coding", "owasp", "sdlc", "application security", "software"]):
        procedures.extend([
            "1. Request secure development standards/guidelines.",
            "2. Verify alignment with OWASP Top 10.",
            "3. Check security testing integration in SDLC (SAST, DAST, SCA).",
            "4. Review most recent application security test results."
        ])
        evidence.extend(["Secure coding standards", "OWASP mapping", "Security test reports", "SDLC documentation"])

    # Secure configuration
    elif any(w in text for w in ["configuration", "hardening", "baseline", "cis", "stig"]):
        procedures.extend([
            "1. Request secure configuration baselines.",
            "2. Verify alignment with CIS benchmarks or equivalent.",
            "3. Check configuration compliance monitoring (drift detection).",
            "4. Review remediation of configuration deviations."
        ])
        evidence.extend(["Configuration baselines", "CIS benchmark mapping", "Compliance scan results", "Deviation remediation records"])

    # Generic fallback
    else:
        procedures.extend([
            "1. Request documentation evidencing compliance with this requirement.",
            "2. Verify the documented process/control is implemented as described.",
            "3. Confirm ongoing compliance through review/monitoring records.",
            "4. Check for any gaps or deviations and note in observations."
        ])
        evidence.extend(["Relevant policy/procedure", "Implementation evidence", "Review/monitoring records"])

    return '\n'.join(procedures), '; '.join(evidence)


def get_method(req_text):
    """Determine the primary assessment method based on requirement text."""
    text = req_text.lower()
    if any(w in text for w in ["monitor", "track", "log", "scan", "test"]):
        return "Inspection"
    elif any(w in text for w in ["communicat", "training", "awareness"]):
        return "Inspection / Inquiry"
    elif any(w in text for w in ["physical", "cctv", "perimeter"]):
        return "Inspection / Observation"
    elif any(w in text for w in ["conduct", "perform", "audit", "assess"]):
        return "Inspection"
    else:
        return "Inspection"


# ── Sheet 1: Methodology & Approach ──────────────────────────────

def create_methodology_sheet(wb):
    ws = wb.active
    ws.title = 'Methodology & Approach'
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    for col in 'CDEFGHIJKLMN':
        ws.column_dimensions[col].width = 12

    r = 1
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='NACSA CoP COMPLIANCE ASSESSMENT — METHODOLOGY & APPROACH').font = TITLE_FONT
    r += 1
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='Cyber Security Act 2024 (Act 854) — NCII Entity Compliance Assessment').font = SUBTITLE_FONT
    r += 2

    # Engagement details
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='ENGAGEMENT DETAILS').font = SUBTITLE_FONT
    r += 1
    details = [
        ('Entity Name:', '[Entity Name]'),
        ('NCII Sector:', '[Sector]'),
        ('Sector Lead:', '[Sector Lead]'),
        ('Engagement Type:', 'Compliance Assessment (Mandatory) + Risk-Based Assessment (Mandatory)'),
        ('Regulatory Basis:', 'Cyber Security Act 2024 (Act 854) — CoP Template v1.6'),
        ('Assessment Date:', '[DD/MM/YYYY]'),
        ('Assessment Period:', '[Start Date] to [End Date]'),
        ('Lead Auditor:', '[Name, Qualification]'),
        ('Engagement Partner:', '[Name, Qualification]'),
        ('Legal Practitioner (Mode A):', '[Name / Firm]'),
        ('Engagement Reference:', '[Ref Number]'),
        ('Report Classification:', 'Sulit (Confidential)'),
    ]
    for label, value in details:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=2, value=value).font = VALUE_FONT
        r += 1
    r += 1

    # Scope
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='SCOPE OF ASSESSMENT').font = SUBTITLE_FONT
    r += 1
    scope_text = (
        'This assessment covers compliance with the NACSA Code of Practice Template v1.6 across 17 control domains '
        '(CoP 4.0-20.0), 126 controls, and 228 testable requirements (210 Tier 1 SHALL + 18 Tier 2 SHOULD/MAY). '
        'Additionally, 42 directive-only requirements from CE Directives 01-10 are assessed separately.'
    )
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value=scope_text).font = VALUE_FONT
    ws.cell(row=r, column=1).alignment = WRAP_ALIGN
    r += 2

    # Assessment Approaches (6 approaches: 2 mandatory + 4 optional)
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='ASSESSMENT APPROACHES').font = SUBTITLE_FONT
    r += 1
    approaches = [
        ('Compliance Assessment (MANDATORY)', 'Assess compliance with CoP Template clauses. Rating: P/SP/TP/TB per control domain.'),
        ('Risk-Based Assessment (MANDATORY)', 'Evaluate risk management processes per CoP 9.0 and annual RA requirement under P.U.(A) 219/2024.'),
        ('Control-Based Assessment (OPTIONAL)', 'Evaluate operating effectiveness of selected controls. Rating: Berkesan/Sebahagian Berkesan/Tidak Berkesan.'),
        ('Technical Testing (OPTIONAL)', 'Vulnerability scanning, penetration testing, configuration review of NCII systems.'),
        ('Physical Inspection (OPTIONAL)', 'On-site physical security inspection of data centres and critical infrastructure housing NCII.'),
        ('Continuous Improvement Review (OPTIONAL)', 'Evaluate maturity trajectory and improvement plans. No compliance rating — qualitative narrative only.'),
    ]
    for label, desc in approaches:
        ws.merge_cells(f'A{r}:N{r}')
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        r += 1
        ws.merge_cells(f'A{r}:N{r}')
        ws.cell(row=r, column=1, value=desc).font = VALUE_FONT
        ws.cell(row=r, column=1).alignment = WRAP_ALIGN
        r += 1
    r += 1

    # Assessment levels
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='ASSESSMENT LEVELS').font = SUBTITLE_FONT
    r += 1
    levels = [
        ('ORG', 'Organisation-level. Assessed ONCE per engagement. Policies, governance, frameworks that apply across all NCII systems.'),
        ('SYSTEM', 'NCII System-level. Assessed PER NCII SYSTEM in scope using sampling. Covers system-specific configurations, controls, and technical measures.'),
    ]
    for lvl, desc in levels:
        ws.cell(row=r, column=1, value=lvl).font = LABEL_FONT
        ws.cell(row=r, column=2, value=desc).font = VALUE_FONT
        ws.cell(row=r, column=2).alignment = WRAP_ALIGN
        r += 1
    r += 1

    # Assessment methods
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='ASSESSMENT METHODS').font = SUBTITLE_FONT
    r += 1
    for method, desc in [
        ('Inspection', 'Examination of documents, records, policies, configurations, and tangible evidence.'),
        ('Inquiry', 'Interviews with management, process owners, and staff.'),
        ('Observation', 'Direct observation of processes, systems, and activities.'),
        ('Confirmation', 'Verification with external or independent parties.'),
        ('Re-performance', 'Auditor independently re-performs the control procedure.'),
    ]:
        ws.cell(row=r, column=1, value=method).font = LABEL_FONT
        ws.cell(row=r, column=2, value=desc).font = VALUE_FONT
        ws.cell(row=r, column=2).alignment = WRAP_ALIGN
        r += 1
    r += 1

    # Evidence hierarchy
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='EVIDENCE HIERARCHY (prefer higher-ranked)').font = SUBTITLE_FONT
    r += 1
    for rank, desc in [
        ('Rank 1', 'Direct observation — auditor directly observes control in operation'),
        ('Rank 2', 'Independent confirmation — third-party evidence (certifications, audit reports)'),
        ('Rank 3', 'System-generated — logs, configurations without manual intervention'),
        ('Rank 4', 'Re-performance — auditor re-performs control procedure'),
        ('Rank 5', 'Documentary — policies, procedures, meeting minutes'),
        ('Rank 6', 'Inquiry — verbal representations from entity personnel'),
    ]:
        ws.cell(row=r, column=1, value=rank).font = LABEL_FONT
        ws.cell(row=r, column=2, value=desc).font = VALUE_FONT
        r += 1
    r += 1

    # Compliance rating scale
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='COMPLIANCE RATING SCALE').font = SUBTITLE_FONT
    r += 1
    for code, desc in [
        ('P — Patuh (Compliant)', 'Requirement fully met: control exists, approved, applied, current, evidenced.'),
        ('SP — Separa Patuh (Partially Compliant)', 'Some elements exist but gaps remain; started but not completed; covers some but not all NCII systems.'),
        ('TP — Tidak Patuh (Non-Compliant)', 'Not met or no evidence; entity not aware of requirement; fundamentally inadequate.'),
        ('TB — Tidak Berkenaan (Not Applicable)', 'Justified exclusion — entity must provide written justification.'),
    ]:
        ws.cell(row=r, column=1, value=code).font = LABEL_FONT
        ws.cell(row=r, column=2, value=desc).font = VALUE_FONT
        ws.cell(row=r, column=2).alignment = WRAP_ALIGN
        r += 1
    r += 1

    # Tier classification
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='TIER CLASSIFICATION').font = SUBTITLE_FONT
    r += 1
    for code, desc in [
        ('Tier 1 (T1) — SHALL clauses', 'Mandatory compliance. Rated P/SP/TP/TB. Non-compliance = potential finding.'),
        ('Tier 2 (T2) — SHOULD/MAY clauses', 'Improvement opportunity. Noted but NOT rated. No findings raised.'),
    ]:
        ws.cell(row=r, column=1, value=code).font = LABEL_FONT
        ws.cell(row=r, column=2, value=desc).font = VALUE_FONT
        ws.cell(row=r, column=2).alignment = WRAP_ALIGN
        r += 1
    r += 1

    # Sampling methodology
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='SAMPLING METHODOLOGY').font = SUBTITLE_FONT
    r += 1
    ws.cell(row=r, column=1, value='Sampling Approach:').font = LABEL_FONT
    ws.merge_cells(f'B{r}:N{r}')
    ws.cell(row=r, column=2, value='Judgmental (not statistical) — appropriate for compliance and limited assurance engagements under Act 854.').font = VALUE_FONT
    ws.cell(row=r, column=2).alignment = WRAP_ALIGN
    r += 2

    ws.cell(row=r, column=1, value='Sample Size Table:').font = LABEL_FONT
    r += 1
    for col, h in enumerate(['Population Size', 'Standard Risk', 'High Risk', 'Rationale'], 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = LABEL_FONT
        c.border = THIN_BORDER
    r += 1
    for pop, std, high, rationale in [
        ('1-5', 'All', 'All', 'Full coverage feasible'),
        ('6-15', '3', '5', '~25-40% coverage'),
        ('16-50', '5', '8', 'Sufficient for pattern detection'),
        ('51-100', '8', '12', 'Diminishing returns beyond this'),
        ('100+', '10', '15', 'Cap with stratification'),
    ]:
        for col, val in enumerate([pop, std, high, rationale], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = VALUE_FONT
            c.border = THIN_BORDER
            c.alignment = WRAP_ALIGN
        r += 1
    r += 1

    ws.cell(row=r, column=1, value='Time-Based Evidence Coverage:').font = LABEL_FONT
    r += 1
    for col, h in enumerate(['Evidence Type', 'Coverage Period', 'Rationale'], 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = LABEL_FONT
        c.border = THIN_BORDER
    r += 1
    for etype, period, rationale in [
        ('Board/committee reports', 'Last 4 quarters', 'Full annual governance cycle'),
        ('Operational records (incidents, changes, patches)', 'Last 3-6 months', 'Recent operating effectiveness'),
        ('Annual processes (risk assessment, DR test, pentest)', 'Last 12 months', 'Full cycle'),
        ('Continuous monitoring (SOC, logs, alerts)', 'Last 30 days', 'Current state verification'),
        ('Policies and frameworks', 'Current version + prior', 'Change tracking'),
    ]:
        for col, val in enumerate([etype, period, rationale], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = VALUE_FONT
            c.border = THIN_BORDER
            c.alignment = WRAP_ALIGN
        r += 1
    r += 1

    # Limitations
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='LIMITATIONS').font = SUBTITLE_FONT
    r += 1
    for lim in [
        '1. Limited assurance engagement — conclusions based on procedures performed, not absolute assurance.',
        '2. Point-in-time (design adequacy) or period assessment (operating effectiveness).',
        '3. Reliance on management representations and documentation provided.',
        '4. Scope guided by Cyber Security Act 2024 (Act 854) and CoP Template v1.6.',
        '5. Does not constitute legal or regulatory advice.',
        '6. The auditor remains solely responsible for the sufficiency of work performed.',
    ]:
        ws.merge_cells(f'A{r}:N{r}')
        ws.cell(row=r, column=1, value=lim).font = VALUE_FONT
        ws.cell(row=r, column=1).alignment = WRAP_ALIGN
        r += 1
    r += 1

    # Sign-off
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value='SIGN-OFF').font = SUBTITLE_FONT
    r += 1
    for s in ['Lead Auditor:', 'Engagement Partner:', 'Entity Representative:']:
        ws.cell(row=r, column=1, value=s).font = LABEL_FONT
        ws.cell(row=r, column=2, value='________________________').font = VALUE_FONT
        ws.cell(row=r, column=6, value='Date:').font = LABEL_FONT
        ws.cell(row=r, column=7, value='______________').font = VALUE_FONT
        r += 1

    r += 1
    ws.merge_cells(f'A{r}:N{r}')
    ws.cell(row=r, column=1, value=(
        f"Generated: {date.today().isoformat()} | Library v{library_data['_meta']['version']} "
        f"(CoP Template v1.6 verbatim clauses) | AWP v5.0 (14-column IESP pattern)"
    )).font = SMALL_FONT

    return ws


# ── Sheet 2: Scoping ─────────────────────────────────────────────

def create_scoping_sheet(wb):
    ws = wb.create_sheet(title='Scoping')
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 20

    r = 1
    ws.merge_cells(f'A{r}:I{r}')
    ws.cell(row=r, column=1, value='ENGAGEMENT SCOPING').font = TITLE_FONT
    r += 2

    # Assessment levels
    ws.merge_cells(f'A{r}:I{r}')
    ws.cell(row=r, column=1, value='ASSESSMENT LEVELS').font = SUBTITLE_FONT
    r += 1
    for level, desc in [
        ('ORG', 'Organisation-level. Assessed ONCE per engagement. Policies, governance, frameworks that apply across all NCII systems.'),
        ('SYSTEM', 'NCII System-level. Assessed PER NCII SYSTEM in scope. System-specific configurations and technical controls.'),
    ]:
        ws.cell(row=r, column=1, value=level).font = LABEL_FONT
        ws.merge_cells(f'B{r}:I{r}')
        ws.cell(row=r, column=2, value=desc).font = VALUE_FONT
        ws.cell(row=r, column=2).alignment = WRAP_ALIGN
        r += 1
    r += 1

    # NCII System Register
    ws.merge_cells(f'A{r}:I{r}')
    ws.cell(row=r, column=1, value='NCII SYSTEM REGISTER').font = SUBTITLE_FONT
    r += 1
    ws.merge_cells(f'A{r}:I{r}')
    ws.cell(row=r, column=1, value='Populate from entity risk register. This drives all system-level testing in domain worksheets.').font = SMALL_FONT
    r += 1

    headers = ['#', 'System Name', 'Description', 'Classification', 'Location',
               'Owner', 'In Risk Register?', 'In Audit Scope?', 'Notes']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
    r += 1

    for i, sys_info in enumerate(NCII_SYSTEMS, 1):
        vals = [str(i), sys_info["name"], sys_info["description"],
                "[Critical/High/Medium]", "[Location]", "[Owner]", "Y/N", "Y/N", ""]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = VALUE_FONT
            c.border = THIN_BORDER
        r += 1
    for _ in range(7):
        for col in range(1, 10):
            ws.cell(row=r, column=col).border = THIN_BORDER
        r += 1
    r += 1

    # Sampling approach
    ws.merge_cells(f'A{r}:I{r}')
    ws.cell(row=r, column=1, value='SAMPLING APPROACH').font = SUBTITLE_FONT
    r += 1
    samples = [
        ('Assessment Period:', '[Start Date] to [End Date]'),
        ('NCII Systems in Scope:', '[From register above]'),
        ('System-Level Sampling:', '1-5: Full population | 6-15: All critical + sample | 16+: Risk-based sample'),
        ('User Access Samples:', '[e.g., 20 user accounts, 10 leavers/transfers]'),
        ('Change Management Samples:', '[e.g., 10 recent changes]'),
        ('Incident Samples:', '[e.g., 15 SOC alerts, all significant incidents in period]'),
    ]
    for label, val in samples:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.merge_cells(f'B{r}:I{r}')
        ws.cell(row=r, column=2, value=val).font = VALUE_FONT
        r += 1

    return ws


# ── Sheet 3: Planning ────────────────────────────────────────────

def create_planning_sheet(wb):
    ws = wb.create_sheet(title='Planning')
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30

    r = 1
    ws.merge_cells(f'A{r}:E{r}')
    ws.cell(row=r, column=1, value='PRE-ENGAGEMENT PLANNING CHECKLIST').font = TITLE_FONT
    r += 1
    ws.merge_cells(f'A{r}:E{r}')
    ws.cell(row=r, column=1, value='Per CE Directive 08 audit process requirements.').font = SMALL_FONT
    r += 2

    headers = ['Step', 'Activity', 'Responsible', 'Status', 'Notes']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
    r += 1

    steps = [
        ('P-01', 'Confirm auditor appointment approved by CE NACSA (CED-08)'),
        ('P-02', 'Sign NDA with NACSA (CED-08 requirement)'),
        ('P-03', 'Issue engagement letter with scope, timeline, and deliverables'),
        ('P-04', 'Confirm audit team qualifications and independence'),
        ('P-05', 'Issue Pre-Engagement Document Request List (PBC) to entity'),
        ('P-06', 'Obtain and review prior audit/RA reports and remediation status'),
        ('P-07', 'Identify NCII systems in scope and complete scoping sheet'),
        ('P-08', 'Define assessment period and sampling approach'),
        ('P-09', 'Determine assessment approaches elected (2 mandatory + optional)'),
        ('P-10', 'Schedule on-site visits and remote assessment sessions'),
        ('P-11', 'Conduct kick-off meeting with entity management'),
        ('P-12', 'Receive and catalogue PBC responses'),
        ('P-13', 'Finalise scoping memorandum'),
        ('P-14', 'Confirm legal practitioner engagement (Mode A/B)'),
    ]
    for ref, activity in steps:
        data = [ref, activity, '', '', '']
        for col, val in enumerate(data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = BODY_FONT
            c.alignment = WRAP_ALIGN
            c.border = THIN_BORDER
        r += 1

    return ws


# ── Sheet 4: Statutory & Legal ───────────────────────────────────

def create_statutory_legal_sheet(wb):
    ws = wb.create_sheet(title='Statutory & Legal')
    set_col_widths(ws, [10, 28, 40, 12, 22, 12, 35, 12, 30, 30])

    # --- Section A: Statutory Obligations ---
    r = 1
    ws.merge_cells("A1:J1")
    ws["A1"] = "STATUTORY & LEGAL OVERLAY — Act 854"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = LEGAL_FILL

    r = 3
    ws.merge_cells(f"A{r}:J{r}")
    ws[f"A{r}"] = "SECTION A: STATUTORY OBLIGATIONS (s17-s26)"
    ws[f"A{r}"].font = SUBTITLE_FONT
    r += 1

    headers = [
        "Section", "Obligation", "Requirement Summary",
        "Penalty Ref", "Penalty", "Compoundable?",
        "Expected Evidence", "Status\n(Met/Not Met)",
        "Tech Auditor Notes", "Legal Practitioner Notes"
    ]
    write_row(ws, r, headers, font=LEGAL_HEADER_FONT, fill=LEGAL_FILL, alignment=CENTER_ALIGN)
    r += 1

    for ob in STATUTORY_OBLIGATIONS:
        section, name, summary, penalty_ref, penalty, compound, evidence = ob
        write_row(ws, r, [section, name, summary, penalty_ref, penalty, compound, evidence, "", "", ""])
        if compound == "No":
            for col in range(1, 11):
                ws.cell(row=r, column=col).fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        r += 1

    # --- Section B: Director Liability (s52) ---
    r += 2
    ws.merge_cells(f"A{r}:J{r}")
    ws[f"A{r}"] = "SECTION B: DIRECTOR & OFFICER LIABILITY (s52)"
    ws[f"A{r}"].font = SUBTITLE_FONT
    r += 1
    ws.merge_cells(f"A{r}:J{r}")
    ws[f"A{r}"] = (
        "s52 creates personal CRIMINAL liability for directors/officers via consent, connivance, or NEGLECT. "
        "Reverse burden of proof — officers must prove due diligence."
    )
    ws[f"A{r}"].font = SMALL_FONT
    ws[f"A{r}"].alignment = WRAP_ALIGN
    r += 2

    # Liability triggers
    for trigger, desc in [
        ("CONSENT", "Officer knew about and approved the breach"),
        ("CONNIVANCE", "Officer was aware of the breach but did nothing to stop it"),
        ("NEGLECT", "Officer failed to exercise proper oversight — NO knowledge/intent required"),
    ]:
        write_row(ws, r, [trigger, desc, "", "", "", "", "", "", "", ""])
        ws.cell(row=r, column=1).font = LABEL_FONT
        r += 1

    r += 1
    s52_headers = [
        "Assessment Area", "Key Questions", "Expected Evidence",
        "", "Assessment\n(Strong/Adequate/Weak)", "",
        "", "", "Assessor Notes", ""
    ]
    write_row(ws, r, s52_headers, font=LEGAL_HEADER_FONT, fill=LEGAL_FILL, alignment=CENTER_ALIGN)
    r += 1

    for area, questions, evidence in S52_ASSESSMENT:
        write_row(ws, r, [area, questions, evidence, "", "", "", "", "", "", ""])
        r += 1

    # --- Section C: Penalty Exposure Matrix ---
    r += 2
    ws.merge_cells(f"A{r}:J{r}")
    ws[f"A{r}"] = "SECTION C: PENALTY EXPOSURE MATRIX"
    ws[f"A{r}"].font = SUBTITLE_FONT
    r += 1

    pen_headers = [
        "Section", "Offence", "Individual Fine", "Imprisonment",
        "Corporate Fine (10x)", "Compoundable?", "Compound Amount",
        "", "Current Exposure\n(High/Med/Low/None)", "Linked Findings"
    ]
    write_row(ws, r, pen_headers, font=LEGAL_HEADER_FONT, fill=LEGAL_FILL, alignment=CENTER_ALIGN)
    r += 1

    for p in PENALTY_MATRIX:
        section, offence, fine, prison, corp_fine, compound, compound_amt = p
        write_row(ws, r, [section, offence, fine, prison, corp_fine, compound, compound_amt, "", "", ""])
        if compound == "No":
            for col in range(1, 11):
                ws.cell(row=r, column=col).fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        r += 1

    return ws


# ── Sheet 5: CE Directive Compliance ─────────────────────────────

def create_directive_compliance_sheet(wb):
    ws = wb.create_sheet("CE Directive Compliance")
    set_col_widths(ws, [12, 30, 55, 14, 18, 35])
    max_col = 6

    ws.merge_cells(f"A1:{get_column_letter(max_col)}1")
    ws["A1"] = "CE DIRECTIVE COMPLIANCE — 42 Mandatory Obligations Not in CoP"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = DIRECTIVE_HEADER_FILL

    ws.merge_cells(f"A2:{get_column_letter(max_col)}2")
    ws["A2"] = (
        "42 Tier 1 requirements from CE Directives 01-10 that are NOT covered by CoP clauses. "
        "Assessed separately from domain worksheets."
    )
    ws["A2"].font = SMALL_FONT
    ws["A2"].alignment = WRAP_ALIGN

    r = 4
    directive_groups = {}
    for req in DIRECTIVE_REQUIREMENTS:
        d = req["ceDirective"]
        if d not in directive_groups:
            directive_groups[d] = []
        directive_groups[d].append(req)

    total_reqs = 0
    for dir_num in sorted(directive_groups.keys()):
        reqs = directive_groups[dir_num]
        dir_name = DIRECTIVE_NAMES.get(dir_num, f"Directive {dir_num}")
        ws.merge_cells(f"A{r}:{get_column_letter(max_col)}{r}")
        ws[f"A{r}"] = f"{dir_name} ({len(reqs)} requirements)"
        ws[f"A{r}"].font = Font(name="Calibri", size=11, bold=True)
        ws[f"A{r}"].fill = DIRECTIVE_FILL
        ws[f"A{r}"].border = THIN_BORDER
        r += 1

        headers = ["Directive", "Para / Source", "Requirement", "Status\n(Met/Not Met)", "Evidence Ref", "Notes"]
        write_row(ws, r, headers, font=DIRECTIVE_HEADER_FONT, fill=DIRECTIVE_HEADER_FILL, alignment=CENTER_ALIGN)
        r += 1

        start_row = r
        for req in reqs:
            source = req.get("verbatimSource", f"CE Directive {dir_num}")
            write_row(ws, r, [
                f"CED-{dir_num}", source, req["text"], "", "",
                f"Act ref: {req.get('actSection', '')}" if req.get("actSection") else "",
            ])
            total_reqs += 1
            r += 1

        if start_row <= r - 1:
            rng = f"D{start_row}:D{r - 1}"
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="equal", formula=['"Met"'], fill=COMPLIANT_FILL)
            )
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="equal", formula=['"Not Met"'], fill=NON_COMPLIANT_FILL)
            )
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:{get_column_letter(max_col)}{r}")
    ws[f"A{r}"] = f"Total directive requirements: {total_reqs}"
    ws[f"A{r}"].font = LABEL_FONT

    return total_reqs


# ── Domain Worksheets (14-column layout) ─────────────────────────

def create_domain_worksheet(wb, domain, sheet_num):
    """Create a domain worksheet using the IESP 14-column pattern."""
    cop = domain["copSection"]
    name = domain["name"]
    slug = domain["slug"]
    sheet_title = f"CoP {cop} {name}"

    # Excel sheet name limit: 31 chars
    if len(sheet_title) > 31:
        sheet_title = sheet_title[:31]

    ws = wb.create_sheet(sheet_title)
    apply_col_widths(ws)

    domain_controls = CONTROLS_BY_DOMAIN.get(slug, [])
    org_controls = [c for c in domain_controls if c["scope"] == "org"]
    sys_controls = [c for c in domain_controls if c["scope"] == "sys"]

    total_t1 = sum(c["tier1Count"] for c in domain_controls)
    total_t2 = sum(c["tier2Count"] for c in domain_controls)

    # Row 1: header row
    r = 1
    write_header_row(ws, r)
    r += 1

    # Track domain rows for scoring
    domain_row_refs = []
    sub_item_idx = 0

    # --- ORG-level controls ---
    if org_controls:
        write_section_row(ws, r, f"CoP {cop} — {name} (Organisation-Level)")
        r += 1

        # Context row
        context = DOMAIN_CONTEXT.get(slug, "")
        if context:
            write_context_row(ws, r, context)
            r += 1

        for ctrl_idx, ctrl in enumerate(org_controls):
            subsection = ctrl["copSubsection"]
            ctrl_name = ctrl["name"]
            awp_ref = f"NACSA-{subsection}"

            # Summary for domain row
            t1_reqs = [req for req in ctrl["requirements"] if req["tier"] == 1]
            t2_reqs = [req for req in ctrl["requirements"] if req["tier"] == 2]
            summary = f"Assess {len(t1_reqs)} mandatory requirement(s)"
            if t2_reqs:
                summary += f" + {len(t2_reqs)} improvement opportunity(ies)"

            # Write domain row
            write_domain_row(ws, r, subsection, awp_ref, 'ORG', ctrl_name, summary)
            domain_row_refs.append(r)
            r += 1

            # Write sub-item rows for each requirement
            for req_idx, req in enumerate(ctrl["requirements"]):
                clause_num = req["clauseNumber"]
                is_tier2 = req["tier"] == 2
                sub_awp_ref = f"NACSA-{subsection}{chr(97 + req_idx)}"

                if is_tier2:
                    sub_proc = f"(Tier 2 -- Improvement Opportunity) {req['text'][:80]}..."
                else:
                    sub_proc = req["text"][:120] + ("..." if len(req["text"]) > 120 else "")

                procedures, evidence = get_assessment_procedures(ctrl, req)
                method = get_method(req["text"])

                write_subitem_row(ws, r, clause_num, sub_awp_ref, sub_proc,
                                  procedures, evidence, method,
                                  is_tier2=is_tier2, alt_row=(sub_item_idx % 2 == 1))
                sub_item_idx += 1
                r += 1

    # --- SYSTEM-level controls ---
    if sys_controls:
        if org_controls:
            r += 1  # gap

        write_section_row(ws, r, f"CoP {cop} — {name} (System-Level: per NCII system)")
        r += 1

        # Context row for sys section if no org context was written
        if not org_controls:
            context = DOMAIN_CONTEXT.get(slug, "")
            if context:
                write_context_row(ws, r, context)
                r += 1

        for ctrl in sys_controls:
            subsection = ctrl["copSubsection"]
            ctrl_name = ctrl["name"]

            for sys_info in NCII_SYSTEMS:
                sys_name = sys_info["name"]
                awp_ref = f"NACSA-{subsection}-{sys_info['id']}"

                t1_reqs = [req for req in ctrl["requirements"] if req["tier"] == 1]
                t2_reqs = [req for req in ctrl["requirements"] if req["tier"] == 2]
                summary = f"{sys_name}: Assess {len(t1_reqs)} mandatory requirement(s)"
                if t2_reqs:
                    summary += f" + {len(t2_reqs)} improvement opportunity(ies)"

                write_domain_row(ws, r, subsection, awp_ref, 'SYSTEM', ctrl_name, summary)
                # Override fill to system colour
                for col in range(1, NUM_COLS + 1):
                    ws.cell(row=r, column=col).fill = SYS_FILL
                ws.cell(row=r, column=13).fill = CONCLUSION_FILL
                ws.cell(row=r, column=14).fill = CONCLUSION_FILL
                domain_row_refs.append(r)
                r += 1

                for req_idx, req in enumerate(ctrl["requirements"]):
                    clause_num = req["clauseNumber"]
                    is_tier2 = req["tier"] == 2
                    sub_awp_ref = f"NACSA-{subsection}{chr(97 + req_idx)}-{sys_info['id']}"

                    if is_tier2:
                        sub_proc = f"(Tier 2 -- Improvement Opportunity) {req['text'][:80]}..."
                    else:
                        sub_proc = f"{sys_name}: {req['text'][:100]}{'...' if len(req['text']) > 100 else ''}"

                    procedures, evidence = get_assessment_procedures(ctrl, req)
                    method = get_method(req["text"])

                    write_subitem_row(ws, r, clause_num, sub_awp_ref, sub_proc,
                                      procedures, evidence, method,
                                      is_tier2=is_tier2, alt_row=(sub_item_idx % 2 == 1))
                    sub_item_idx += 1
                    r += 1

    # Add conditional formatting for conclusion column (M)
    if domain_row_refs:
        min_row = min(domain_row_refs)
        max_row = r
        rng = f"M{min_row}:M{max_row}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"P"'], fill=COMPLIANT_FILL)
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"SP"'], fill=PARTIAL_FILL)
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"TP"'], fill=NON_COMPLIANT_FILL)
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"TB"'], fill=NA_FILL)
        )

    return {
        "t1": total_t1, "t2": total_t2,
        "org": len(org_controls), "sys": len(sys_controls),
        "controls": len(domain_controls),
        "domain_rows": len(domain_row_refs),
        "sheet_title": sheet_title,
    }


# ── Finding Register ─────────────────────────────────────────────

def create_finding_register(wb):
    ws = wb.create_sheet("Finding Register")
    set_col_widths(ws, [14, 30, 14, 14, 14, 40, 35, 30, 35, 35, 14, 18, 16, 18, 14])

    ws.merge_cells("A1:O1")
    ws["A1"] = "FINDING REGISTER — 5 C's Framework with Legal Classification"
    ws["A1"].font = TITLE_FONT

    r = 3
    headers = [
        "Finding ID", "Title", "CoP Domain", "CoP Ref", "System\n(if applicable)",
        "Condition\n(Keadaan Sebenar)", "Criteria\n(Act/CoP Reference)",
        "Cause\n(Punca Masalah)", "Consequence\n(Impak + Penalty Exposure)",
        "Corrective Action\n(Peluang Penambahbaikan)",
        "Severity\n(K/T/S/R)", "Legal Classification",
        "Compoundable?", "Status\n(Open/In Progress/Remediated/Closed)",
        "Target Date"
    ]
    write_row(ws, r, headers, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN)
    r += 1

    for i in range(1, 31):
        finding_id = f"CoP-{date.today().year}-{i:03d}"
        write_row(ws, r, [finding_id] + [""] * 14)
        r += 1

    r += 2
    ws.merge_cells(f"A{r}:O{r}")
    ws[f"A{r}"] = "Severity: K = Kritikal (30 days) | T = Tinggi (90 days) | S = Sederhana (180 days) | R = Rendah (next audit cycle)"
    ws[f"A{r}"].font = SMALL_FONT
    r += 1
    ws.merge_cells(f"A{r}:O{r}")
    ws[f"A{r}"] = "Legal Classification: Statutory (s17-s26) | CoP (Code of Practice) | Standard (CE standard) | Improvement (not a finding)"
    ws[f"A{r}"].font = SMALL_FONT

    return ws


# ── Scoring Dashboard ────────────────────────────────────────────

def create_scoring_dashboard(wb, domain_sheet_names):
    ws = wb.create_sheet(title='Scoring Dashboard')
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 18

    r = 1
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(row=r, column=1, value='SCORING DASHBOARD').font = TITLE_FONT
    r += 2

    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(row=r, column=1,
            value='Conclusions are at the CONTROL DOMAIN level only (domain rows with green/orange fill). '
                  'Sub-item rows have observations but no conclusion. '
                  'Compliance ratings: P = Patuh | SP = Separa Patuh | TP = Tidak Patuh | TB = Tidak Berkenaan.').font = CONTEXT_FONT
    ws.cell(row=r, column=1).alignment = WRAP_ALIGN
    r += 2

    # Overall Summary
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(row=r, column=1, value='OVERALL SUMMARY').font = SUBTITLE_FONT
    r += 1

    for col, h in enumerate(['Metric', 'Value'], 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
    r += 1

    overall_start_row = r
    for label in [
        'Total Patuh (P)',
        'Total Separa Patuh (SP)',
        'Total Tidak Patuh (TP)',
        'Total Tidak Berkenaan (TB)',
        'Total Not Assessed',
        'Total Control Domains',
        'Overall Compliance %',
    ]:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2).font = VALUE_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        r += 1

    r += 1
    # Supplementary Opinion Indicator
    ws.merge_cells(f'A{r}:B{r}')
    ws.cell(row=r, column=1, value='SUPPLEMENTARY OPINION INDICATOR').font = SUBTITLE_FONT
    r += 1
    ws.cell(row=r, column=1, value='Opinion').font = LABEL_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=2).border = THIN_BORDER
    opinion_row = r
    r += 1
    ws.cell(row=r, column=1, value='Criteria:').font = CONTEXT_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.merge_cells(f'B{r}:H{r}')
    ws.cell(row=r, column=2,
            value='Memuaskan: TP=0 and K findings=0  |  '
                  'Perlu Penambahbaikan: TP>0 and TP<=5, or T findings>0  |  '
                  'Tidak Memuaskan: TP>5, or K findings>0').font = CONTEXT_FONT
    ws.cell(row=r, column=2).alignment = WRAP_ALIGN
    ws.cell(row=r, column=2).border = THIN_BORDER
    r += 2

    # Per-Domain Breakdown
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(row=r, column=1, value='PER-DOMAIN BREAKDOWN').font = SUBTITLE_FONT
    r += 1

    ps_headers = ['Domain Sheet', 'P (Patuh)', 'SP (Separa Patuh)',
                  'TP (Tidak Patuh)', 'TB', 'Not Assessed', 'Total Domains',
                  'Compliance %']
    for col, h in enumerate(ps_headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = CENTER_ALIGN
    r += 1

    per_sheet_first_row = r

    for sheet_name in domain_sheet_names:
        ws.cell(row=r, column=1, value=sheet_name).font = LABEL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER

        safe = sheet_name.replace("'", "''")

        # P count
        ws.cell(row=r, column=2).font = VALUE_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=2).fill = COMPLIANT_FILL
        ws.cell(row=r, column=2, value=f"=COUNTIF('{safe}'!M:M,\"P\")")

        # SP count
        ws.cell(row=r, column=3).font = VALUE_FONT
        ws.cell(row=r, column=3).border = THIN_BORDER
        ws.cell(row=r, column=3).fill = PARTIAL_FILL
        ws.cell(row=r, column=3, value=f"=COUNTIF('{safe}'!M:M,\"SP\")")

        # TP count
        ws.cell(row=r, column=4).font = VALUE_FONT
        ws.cell(row=r, column=4).border = THIN_BORDER
        ws.cell(row=r, column=4).fill = NON_COMPLIANT_FILL
        ws.cell(row=r, column=4, value=f"=COUNTIF('{safe}'!M:M,\"TP\")")

        # TB count
        ws.cell(row=r, column=5).font = VALUE_FONT
        ws.cell(row=r, column=5).border = THIN_BORDER
        ws.cell(row=r, column=5).fill = NA_FILL
        ws.cell(row=r, column=5, value=f"=COUNTIF('{safe}'!M:M,\"TB\")")

        # Not Assessed: domain rows with non-blank B and C (level set) and blank M
        ws.cell(row=r, column=6).font = VALUE_FONT
        ws.cell(row=r, column=6).border = THIN_BORDER
        ws.cell(row=r, column=6,
                value=f"=COUNTIFS('{safe}'!B:B,\"<>\",'{safe}'!C:C,\"<>\",'{safe}'!M:M,\"=\")"
                      f"-COUNTIF('{safe}'!B:B,\"AWP Ref\")")

        # Total domains (rows with non-blank B and non-blank C, minus header)
        ws.cell(row=r, column=7).font = VALUE_FONT
        ws.cell(row=r, column=7).border = THIN_BORDER
        ws.cell(row=r, column=7,
                value=f"=COUNTIFS('{safe}'!B:B,\"<>\",'{safe}'!C:C,\"<>\")"
                      f"-COUNTIF('{safe}'!B:B,\"AWP Ref\")")

        # Compliance %
        b_col = f'B{r}'
        g_col = f'G{r}'
        e_col = f'E{r}'
        f_col = f'F{r}'
        ws.cell(row=r, column=8).font = VALUE_FONT
        ws.cell(row=r, column=8).border = THIN_BORDER
        ws.cell(row=r, column=8).number_format = '0.0%'
        ws.cell(row=r, column=8,
                value=f'=IF(({g_col}-{e_col}-{f_col})=0,"\u2014",{b_col}/({g_col}-{e_col}-{f_col}))')

        r += 1

    per_sheet_last_row = r - 1

    # Fill Overall Summary formulas
    ps_b = f'B{per_sheet_first_row}:B{per_sheet_last_row}'
    ps_c = f'C{per_sheet_first_row}:C{per_sheet_last_row}'
    ps_d = f'D{per_sheet_first_row}:D{per_sheet_last_row}'
    ps_e = f'E{per_sheet_first_row}:E{per_sheet_last_row}'
    ps_f = f'F{per_sheet_first_row}:F{per_sheet_last_row}'
    ps_g = f'G{per_sheet_first_row}:G{per_sheet_last_row}'

    or_row = overall_start_row
    ws.cell(row=or_row, column=2, value=f'=SUM({ps_b})').fill = COMPLIANT_FILL
    or_row += 1
    ws.cell(row=or_row, column=2, value=f'=SUM({ps_c})').fill = PARTIAL_FILL
    sp_cell = f'B{or_row}'
    or_row += 1
    ws.cell(row=or_row, column=2, value=f'=SUM({ps_d})').fill = NON_COMPLIANT_FILL
    tp_cell = f'B{or_row}'
    or_row += 1
    ws.cell(row=or_row, column=2, value=f'=SUM({ps_e})').fill = NA_FILL
    tb_cell = f'B{or_row}'
    or_row += 1
    ws.cell(row=or_row, column=2, value=f'=SUM({ps_f})')
    na_cell = f'B{or_row}'
    or_row += 1
    ws.cell(row=or_row, column=2, value=f'=SUM({ps_g})')
    total_cell = f'B{or_row}'
    or_row += 1
    p_cell = f'B{overall_start_row}'
    ws.cell(row=or_row, column=2).number_format = '0.0%'
    ws.cell(row=or_row, column=2,
            value=f'=IF(({total_cell}-{tb_cell}-{na_cell})=0,"\u2014",'
                  f'{p_cell}/({total_cell}-{tb_cell}-{na_cell}))')

    # Opinion formula
    ws.cell(row=opinion_row, column=2,
            value=f'=IF({tp_cell}>5,"Tidak Memuaskan (Unsatisfactory)",'
                  f'IF({tp_cell}>0,"Perlu Penambahbaikan (Needs Improvement)",'
                  f'IF({sp_cell}<=3,"Memuaskan (Satisfactory)","Perlu Penambahbaikan (Needs Improvement)")))')

    # Level breakdown
    r += 2
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(row=r, column=1, value='LEVEL BREAKDOWN').font = SUBTITLE_FONT
    r += 1

    lvl_headers = ['Level', 'P (Patuh)', 'SP (Separa Patuh)',
                   'TP (Tidak Patuh)', 'TB', 'Not Assessed', 'Total', '']
    for col, h in enumerate(lvl_headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = CENTER_ALIGN
    r += 1

    for level in ['ORG', 'SYSTEM']:
        ws.cell(row=r, column=1, value=level).font = LABEL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER

        conclusions = [
            ('P', 2, COMPLIANT_FILL),
            ('SP', 3, PARTIAL_FILL),
            ('TP', 4, NON_COMPLIANT_FILL),
            ('TB', 5, NA_FILL),
        ]
        for conc, col_idx, fill in conclusions:
            parts = []
            for sn in domain_sheet_names:
                safe = sn.replace("'", "''")
                parts.append(f"COUNTIFS('{safe}'!C:C,\"{level}\",'{safe}'!M:M,\"{conc}\")")
            formula = '=' + '+'.join(parts) if parts else '=0'
            c = ws.cell(row=r, column=col_idx, value=formula)
            c.font = VALUE_FONT
            c.border = THIN_BORDER
            c.fill = fill

        # Not Assessed by level
        na_parts = []
        for sn in domain_sheet_names:
            safe = sn.replace("'", "''")
            na_parts.append(f"COUNTIFS('{safe}'!C:C,\"{level}\",'{safe}'!M:M,\"=\")")
        c = ws.cell(row=r, column=6, value='=' + '+'.join(na_parts) if na_parts else '=0')
        c.font = VALUE_FONT
        c.border = THIN_BORDER

        ws.cell(row=r, column=7, value=f'=SUM(B{r}:F{r})').font = VALUE_FONT
        ws.cell(row=r, column=7).border = THIN_BORDER

        r += 1

    # Finding summary
    r += 2
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(row=r, column=1, value='FINDING SUMMARY').font = SUBTITLE_FONT
    r += 1

    finding_sheet = 'Finding Register'
    safe_fr = finding_sheet.replace("'", "''")
    finding_summary = [
        ("Critical (Kritikal) findings", f"=COUNTIF('{safe_fr}'!K:K,\"K\")"),
        ("High (Tinggi) findings", f"=COUNTIF('{safe_fr}'!K:K,\"T\")"),
        ("Medium (Sederhana) findings", f"=COUNTIF('{safe_fr}'!K:K,\"S\")"),
        ("Low (Rendah) findings", f"=COUNTIF('{safe_fr}'!K:K,\"R\")"),
    ]
    fs_start = r
    for label, formula in finding_summary:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=formula).font = VALUE_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        r += 1
    ws.cell(row=r, column=1, value="Total findings").font = LABEL_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=2, value=f"=SUM(B{fs_start}:B{r - 1})").font = LABEL_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    r += 2

    # Supplementary Opinion section
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(row=r, column=1, value='SUPPLEMENTARY OPINION (Rumusan)').font = SUBTITLE_FONT
    r += 1
    opinions = [
        ("Memuaskan (Satisfactory)", "No Critical/High findings. TP count zero/near-zero. Active remediation plans."),
        ("Perlu Penambahbaikan (Needs Improvement)", "One or more High findings, OR multiple Medium findings with systemic gaps."),
        ("Tidak Memuaskan (Unsatisfactory)", "One or more Critical findings, OR multiple High findings across core domains."),
    ]
    for opinion, criteria in opinions:
        ws.cell(row=r, column=1, value=opinion).font = LABEL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=criteria).font = VALUE_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=2).alignment = WRAP_ALIGN
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='Selected opinion:').font = LABEL_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=2).border = THIN_BORDER
    r += 2

    # Narrative inputs
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(row=r, column=1, value='NARRATIVE CONCLUSION (Rumusan) INPUTS').font = SUBTITLE_FONT
    r += 1
    for item in [
        "Compliance posture summary (counts, %, strongest/weakest domains)",
        "Risk assessment summary (finding distribution by severity, key risk areas)",
        "Key areas requiring attention and immediate action (CE Directive 08 para 23(j))",
        "CE Directive compliance status (42 directive requirements)",
        "Prior audit follow-up (if applicable)",
        "Legal practitioner input on statutory compliance (Mode A)",
    ]:
        ws.cell(row=r, column=1, value=item).font = BODY_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.merge_cells(f'B{r}:H{r}')
        ws.cell(row=r, column=2).border = THIN_BORDER
        r += 1

    return ws


# ── Document Request (PBC) ───────────────────────────────────────

def create_document_request(wb):
    ws = wb.create_sheet("Document Request")
    set_col_widths(ws, [10, 14, 40, 14, 14, 14, 30])

    ws.merge_cells("A1:G1")
    ws["A1"] = "PRE-ASSESSMENT DOCUMENT REQUEST LIST (PBC)"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:G2")
    ws["A2"] = "Essential items: request 2 weeks before fieldwork. Important items: obtained during fieldwork."
    ws["A2"].font = SMALL_FONT

    r = 4
    headers = [
        "Item #", "CoP Domain", "Document / Evidence", "Priority\n(Essential/Important)",
        "Format", "Status\n(Provided/In Progress/N-A)", "Notes"
    ]
    write_row(ws, r, headers, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN)
    r += 1

    doc_requests = [
        ("4.0", "Governance committee ToR and meeting minutes (12 months)", "Essential"),
        ("4.0", "CISO/ICTSO appointment letter and reporting structure", "Essential"),
        ("4.0", "Board resolutions on cyber security", "Essential"),
        ("5.0", "Approved cyber security policy (current version)", "Essential"),
        ("5.0", "Policy acknowledgement records (Surat Aku Janji)", "Important"),
        ("5.0", "Cyber security objectives register with KPIs", "Important"),
        ("6.0", "Cyber security team org chart and job descriptions", "Essential"),
        ("6.0", "Training plan and completion records", "Essential"),
        ("6.0", "NC4/NACSA/RMP contact register", "Essential"),
        ("7.0", "IT Acceptable Use Policy", "Essential"),
        ("7.0", "Latest audit report and CE submission confirmation", "Essential"),
        ("7.0", "Vendor/supplier register with CSSP licence records", "Essential"),
        ("8.0", "NCII asset register / CMDB extract", "Essential"),
        ("8.0", "Asset classification scheme", "Essential"),
        ("8.0", "Capacity monitoring dashboards", "Important"),
        ("9.0", "Risk assessment report (dated within 12 months)", "Essential"),
        ("9.0", "Risk register and risk treatment plan", "Essential"),
        ("9.0", "CE submission confirmation for RA", "Essential"),
        ("10.0", "Change management procedure", "Essential"),
        ("10.0", "Sample of 10 change records with approvals", "Important"),
        ("11.0", "DLP solution deployment documentation", "Essential"),
        ("11.0", "Data privacy procedure and PDPA alignment mapping", "Essential"),
        ("11.0", "Secure coding standards (OWASP mapping)", "Important"),
        ("11.0", "Secure configuration baselines (CIS/STIG)", "Important"),
        ("12.0", "NDA templates (all types) and signed samples", "Essential"),
        ("13.0", "Physical security procedure", "Essential"),
        ("13.0", "CCTV coverage map and access log samples", "Important"),
        ("14.0", "Network architecture diagrams (physical and logical)", "Essential"),
        ("14.0", "Firewall rule documentation", "Essential"),
        ("14.0", "IDS/IPS and EDR/AV deployment records", "Important"),
        ("15.0", "IAM policy and MFA deployment evidence", "Essential"),
        ("15.0", "Privileged account inventory", "Essential"),
        ("15.0", "Latest user access review records", "Important"),
        ("16.0", "Vulnerability Management Program document", "Essential"),
        ("16.0", "Latest vulnerability scan and penetration test reports", "Essential"),
        ("16.0", "Remediation tracker with SLA compliance", "Important"),
        ("17.0", "Security Monitoring Procedure (SMP)", "Essential"),
        ("17.0", "SIEM log source inventory", "Essential"),
        ("17.0", "MTTD/MTTR KPI dashboards", "Important"),
        ("18.0", "Cyber Security Incident Management Plan (CSIMP)", "Essential"),
        ("18.0", "CSIRT charter and 3 authorised person appointment letters", "Essential"),
        ("18.0", "NC4 portal access evidence", "Essential"),
        ("18.0", "6-hour notification tabletop exercise results", "Important"),
        ("19.0", "BCP document with BIA and RTO/RPO definitions", "Essential"),
        ("19.0", "DR failover test records and results", "Essential"),
        ("19.0", "Backup job logs (90-day analysis)", "Important"),
        ("20.0", "Applicable sector-specific CoP (if published)", "Essential"),
        ("CED", "Authorised person details submitted to CE NACSA (CED-01)", "Essential"),
        ("CED", "NCSB Self-Assessment (CED-04)", "Essential"),
        ("CED", "Auditor appointment approval from NACSA (CED-08)", "Essential"),
        ("CED", "Auditor NDA signed with NACSA (CED-08)", "Essential"),
        ("CED", "National Cyber Crisis Management Plan receipt (CED-10)", "Essential"),
    ]

    item_num = 1
    for domain_ref, doc, priority in doc_requests:
        write_row(ws, r, [f"PBC-{item_num:03d}", domain_ref, doc, priority, "PDF/DOCX", "", ""])
        item_num += 1
        r += 1

    r += 2
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = f"Total document requests: {item_num - 1}"
    ws[f"A{r}"].font = LABEL_FONT

    return ws


# ── Main ─────────────────────────────────────────────────────────

def main():
    wb = Workbook()

    print("Building NACSA CoP Core AWP v5.0 (14-column IESP pattern)...")
    print(f"  Source: controls/library.json v{library_data['_meta']['version']}")
    print(f"  Controls: {library_data['_meta']['controlCount']}")
    print(f"  Requirements: {library_data['_meta']['totalRequirementCount']} "
          f"(T1: {library_data['_meta']['tier1RequirementCount']}, "
          f"T2: {library_data['_meta']['tier2RequirementCount']})")
    print(f"  Directive requirements: {len(DIRECTIVE_REQUIREMENTS)}")
    print()

    # Sheet 1: Methodology & Approach
    create_methodology_sheet(wb)
    print("  [1] Methodology & Approach")

    # Sheet 2: Scoping
    create_scoping_sheet(wb)
    print("  [2] Scoping")

    # Sheet 3: Planning
    create_planning_sheet(wb)
    print("  [3] Planning")

    # Sheet 4: Statutory & Legal
    create_statutory_legal_sheet(wb)
    print("  [4] Statutory & Legal")

    # Sheet 5: CE Directive Compliance
    dir_count = create_directive_compliance_sheet(wb)
    print(f"  [5] CE Directive Compliance ({dir_count} requirements)")

    # Sheets 6-22: Domain Worksheets (17 domains with controls)
    active_domains = [d for d in COP_DOMAINS if d["slug"] in CONTROLS_BY_DOMAIN]
    domain_sheet_names = []

    total_controls = 0
    total_t1 = 0
    total_t2 = 0
    total_domain_rows = 0

    for i, domain in enumerate(active_domains):
        sheet_num = 6 + i
        counts = create_domain_worksheet(wb, domain, sheet_num)
        domain_sheet_names.append(counts["sheet_title"])
        total_controls += counts["controls"]
        total_t1 += counts["t1"]
        total_t2 += counts["t2"]
        total_domain_rows += counts["domain_rows"]

    last_domain_sheet = 6 + len(active_domains) - 1
    print(f"  [6-{last_domain_sheet}] Domain Worksheets ({len(active_domains)} domains, "
          f"{total_controls} controls, {total_t1} T1 + {total_t2} T2 requirements, "
          f"{total_domain_rows} domain conclusion rows)")

    # Sheet: Finding Register
    create_finding_register(wb)
    print(f"  [{last_domain_sheet + 1}] Finding Register")

    # Sheet: Scoring Dashboard
    create_scoring_dashboard(wb, domain_sheet_names)
    print(f"  [{last_domain_sheet + 2}] Scoring Dashboard")

    # Sheet: Document Request
    create_document_request(wb)
    print(f"  [{last_domain_sheet + 3}] Document Request")

    # Save
    output_path = os.path.join(SCRIPT_DIR, "NACSA-CoP-Core-WorkProgram.xlsx")
    wb.save(output_path)
    print(f"\nSaved: {output_path}")

    # Copy to Tech-Audit
    for ta_path in [
        os.path.expanduser("~/claude/Tech-Audit/NACSA"),
        os.path.expanduser("~/claude/tech-audit/NACSA"),
    ]:
        if os.path.isdir(ta_path):
            ta_output = os.path.join(ta_path, "NACSA-CoP-Core-WorkProgram.xlsx")
            wb.save(ta_output)
            print(f"Saved: {ta_output}")
            break
    else:
        print("Note: Tech-Audit/NACSA not found — skipped copy")

    # Print summary
    total_sheets = len(wb.sheetnames)
    print(f"\nWorkbook contains {total_sheets} sheets:")
    for s in wb.sheetnames:
        print(f"  - {s}")

    print(f"\n--- Summary ---")
    print(f"  Domain worksheets: {len(active_domains)}")
    print(f"  Controls: {total_controls}")
    print(f"  CoP requirements: T1={total_t1}, T2={total_t2}, Total={total_t1 + total_t2}")
    print(f"  Directive requirements: {dir_count}")
    print(f"  Domain conclusion rows: {total_domain_rows}")
    print(f"  Total sheets: {total_sheets}")


if __name__ == "__main__":
    main()
